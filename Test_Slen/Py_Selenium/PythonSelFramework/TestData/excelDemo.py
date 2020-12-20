import openpyxl

filename = "C:\\Users\\jsun\\Documents\\Desk_1\\Test_Slen\\Test_Slen\\PythonSelFramework\\TestData\\data.xlsx"
book =openpyxl.load_workbook(filename)
sheet =book.active
Dict = {}
cell =sheet.cell(row=2, column=4)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
book.save(filename)

val = sheet.cell(row=2, column=2).value
# print(f"type(val) is {type(val)}")
# type(val) is <class 'str'>
# print(f"val is {val}")

print(sheet.max_row)

print(sheet.max_column)
#
print(sheet['A3'].value)
#
for i in range(1,sheet.max_row+1):  # to get rows
    if sheet.cell(row =i,column=1).value == "Testcase2":

        for j in range(2,sheet.max_column+1):#to get columns
            #Dict["lastname"]="shetty
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)







